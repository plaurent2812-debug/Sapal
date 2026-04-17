#!/usr/bin/env python3
"""
Indexe l'Excel tarif Procity par référence parent.

Structure Excel (onglet MOBILIER URBAIN / AIRES DE JEUX / ÉQUIPEMENTS SPORTIFS) :
  - Une ligne par (ref_parent, coloris) — pas par variant complet
  - Le coloris "Standard" regroupe tous les RAL hors stock
  - Le stock explicite donne la liste des RAL en stock (ex: "En stock en RAL 3004 / 6005 / 9005 et GPRO")
  - Le prix net est constant pour toutes les variantes d'une même ref parent
  - Le délai dépend : stock => "Disponible sur stock - 2 à 5 jours" / sinon => "3 semaines" (col delai)

Sortie d'index :
    {
      "206200": {
        "category": "AMÉNAGEMENT DE LA RUE",
        "type": "BARRIÈRES DE VILLE",
        "designation": "BARRIÈRE MAIN COURANTE LISBONNE",
        "dimensions": "1000 mm",
        "poids": "11 kg",
        "prix_public": 132.0,
        "prix_net": 92.40,
        "delai_default": "3 semaines",
        "stock_colors": ["3004", "6005", "9005", "GPRO"],
        "delai_stock": "Disponible sur stock - expédition sous 2 à 5 jours maximum",
        "url": "https://procity.eu/fr/...",
        "page": 16,
        "marque": "PROCITY",
        "remise": 0.3
      }
    }
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl


EXCEL_PATH = Path(
    "/Users/pierrelaurent/Desktop/OptiPro/Clients/SAPAL/Fournisseurs/Procity/"
    "tarifprocityvialux2026-fr.v1.7-699.xlsx"
)

SHEETS = {
    "MOBILIER URBAIN": {"ref": 0, "cat": 2, "typ": 3, "des": 4, "dim": 9, "poids": 10,
                        "prix_pub": 11, "prix_net": 12, "coloris": 13, "stock": 14,
                        "delai": 15, "page": 16, "url": 17, "marque": 18, "remise": 19},
    "AIRES DE JEUX": {"ref": 0, "cat": 2, "typ": 3, "des": 4, "dim": 9, "poids": 10,
                      "prix_pub": 7, "prix_net": 8, "coloris": 9, "stock": 10,
                      "delai": 11, "page": 12, "url": 13, "marque": 14, "remise": 15},
    "ÉQUIPEMENTS SPORTIFS": {"ref": 0, "cat": 2, "typ": 3, "des": 4, "dim": 9, "poids": 10,
                             "prix_pub": 7, "prix_net": 8, "coloris": 9, "stock": 10,
                             "delai": 11, "page": 12, "url": 13, "marque": 14, "remise": 15},
}


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------


def extract_stock_colors(stock_text: str) -> list[str]:
    """
    "En stock en RAL 3004 / 6005 / 9005 et GPRO" => ['3004','6005','9005','GPRO']
    """
    if not stock_text or not isinstance(stock_text, str) or stock_text.strip() in {"-", ""}:
        return []
    # Capture les codes RAL à 4 chiffres + noms spéciaux (GPRO, S025, etc.)
    codes = re.findall(r"\b(?:RAL\s*)?(\d{4}|GPRO|S\d{3}|GPROC?)\b", stock_text, re.IGNORECASE)
    # Normaliser
    out = []
    for c in codes:
        c = c.upper().replace("RAL", "").strip()
        if c and c not in out:
            out.append(c)
    return out


def normalize_delai(delai_raw: Any) -> str:
    """Transforme le délai Excel en texte lisible."""
    if delai_raw is None:
        return ""
    s = str(delai_raw).strip()
    if not s:
        return ""
    # Nombres seuls → "X semaines"
    if re.fullmatch(r"\d+(?:[.,]\d+)?", s):
        n = float(s.replace(",", "."))
        if n == int(n):
            n = int(n)
        return f"{n} semaines"
    return s


def num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def format_weight(v: Any) -> str:
    n = num(v)
    if n is None:
        return ""
    if n == int(n):
        return f"{int(n)} kg"
    return f"{n} kg"


# ----------------------------------------------------------------------------
# Build index
# ----------------------------------------------------------------------------


def build_index(excel_path: Path = EXCEL_PATH) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    index: dict[str, dict[str, Any]] = {}

    for sheet_name, cols in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row or not row[cols["ref"]]:
                continue
            ref_raw = row[cols["ref"]]
            ref = str(ref_raw).strip()
            if not ref or ref.lower() in {"reference", "nouveautés"}:
                continue

            url = row[cols["url"]] if len(row) > cols["url"] else None
            if not isinstance(url, str) or not url.startswith("http"):
                url = None

            stock_text = row[cols["stock"]] if len(row) > cols["stock"] else None
            coloris_text = row[cols["coloris"]] if len(row) > cols["coloris"] else None

            stock_colors = extract_stock_colors(stock_text if isinstance(stock_text, str) else "")

            # Ligne "Standard" = porte le délai de fabrication (col delai = nombre = semaines)
            # Les autres lignes coloris (3004, 6005...) sont redondantes, portent juste
            # "Disponible sur stock" dans la col delai.
            # NB: les RAL sont souvent écrits comme int dans Excel (3004, 6005...)
            coloris_str = str(coloris_text).strip() if coloris_text is not None else ""
            is_standard_line = (
                coloris_str.lower() in {"standard", "standard/galva"}
                or coloris_str == ""
                or coloris_str == "-"
            )
            is_stock_line = not is_standard_line

            entry = index.setdefault(ref, {
                "ref": ref,
                "url": url,
                "sheet": sheet_name,
                "category": row[cols["cat"]],
                "type": row[cols["typ"]],
                "designation": row[cols["des"]],
                "dimensions": str(row[cols["dim"]] or "").strip() if cols["dim"] < len(row) else "",
                "poids": format_weight(row[cols["poids"]]) if cols["poids"] < len(row) else "",
                "prix_public": num(row[cols["prix_pub"]]),
                "prix_net": num(row[cols["prix_net"]]),
                "delai_default": "",
                "delai_stock": "",
                "stock_colors": [],
                "page": row[cols["page"]] if cols["page"] < len(row) else None,
                "marque": row[cols["marque"]] if cols["marque"] < len(row) else None,
                "remise": num(row[cols["remise"]]) if cols["remise"] < len(row) else None,
            })
            if not entry["url"] and url:
                entry["url"] = url

            delai = normalize_delai(row[cols["delai"]])
            if is_stock_line:
                # Ligne détail d'un coloris en stock : le délai Excel est "Disponible sur stock..."
                if delai:
                    entry["delai_stock"] = delai
            else:
                # Ligne "Standard" : le délai est celui des coloris hors stock
                if delai:
                    entry["delai_default"] = delai
                if stock_colors:
                    entry["stock_colors"] = stock_colors

    # Nettoyage : délai fallback si une des deux manque
    for entry in index.values():
        if not entry["delai_stock"] and entry["stock_colors"]:
            entry["delai_stock"] = "Disponible sur stock - expédition sous 2 à 5 jours maximum"
        if not entry["delai_default"]:
            # Si le délai "Standard" manque mais qu'on a un délai stock,
            # c'est qu'il n'y a aucune ligne Standard => tous les coloris sont en stock
            entry["delai_default"] = entry["delai_stock"] or "Nous consulter"

    return index


def get_variant_data(index: dict, ref_variant: str) -> dict[str, Any] | None:
    """
    Depuis une ref variant "206200.9005" retourne les données de la ref parent "206200"
    enrichies avec le bon délai pour ce coloris.
    """
    parent_ref = ref_variant.split(".")[0]
    entry = index.get(parent_ref)
    if not entry:
        return None

    # Détermine le délai selon que le coloris (suffix) est en stock ou non
    suffix = ref_variant.split(".", 1)[1] if "." in ref_variant else ""
    suffix_norm = suffix.upper()

    # Match flexible (GPRO vs GPROC, etc.)
    is_in_stock = any(
        sc.upper().replace("C", "") == suffix_norm.replace("C", "")
        or sc.upper() == suffix_norm
        for sc in entry.get("stock_colors", [])
    )
    delai = entry["delai_stock"] if is_in_stock and entry["delai_stock"] else entry["delai_default"]

    return {
        **entry,
        "variant_ref": ref_variant,
        "coloris_code": suffix,
        "delai_resolved": delai,
        "is_in_stock": is_in_stock,
    }


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------


def main() -> int:
    out_path = Path(__file__).parent / "output_preview" / "excel_index.json"
    out_path.parent.mkdir(exist_ok=True)
    idx = build_index()
    out_path.write_text(json.dumps(idx, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"✅ Index: {len(idx)} refs parent indexées")
    print(f"   → {out_path}")

    # Test rapide sur quelques refs Lisbonne
    if len(sys.argv) > 1:
        for q in sys.argv[1:]:
            result = get_variant_data(idx, q)
            print(f"\n— Test '{q}':")
            if not result:
                print("   ✗ pas trouvé")
            else:
                print(f"   ✓ {result['designation']}")
                print(f"     prix_net: {result['prix_net']} € | poids: {result['poids']} | dim: {result['dimensions']}")
                print(f"     stock_colors: {result['stock_colors']}")
                print(f"     in_stock: {result['is_in_stock']} → délai: {result['delai_resolved']}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
