#!/usr/bin/env python3
"""
Procity Variantes Import
========================
1. Lit l'Excel Procity (source prix/delai/poids/ref par ligne)
2. Scrappe chaque URL procity.eu (structure combinaisons + images)
3. Upload les images dans Supabase Storage (bucket: product-images)
4. Upsert product_variants dans Supabase

Usage:
  python3 scrape_and_import.py [--dry-run] [--limit N] [--product-id ID]

Options:
  --dry-run     Affiche ce qui serait insere sans toucher la base
  --limit N     Traite seulement N produits (pour tests)
  --product-id  Traite seulement ce product_id Procity (ex: "206200")
"""

import os
import time
import hashlib
import argparse
import re
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup

# -- Config ------------------------------------------------------------------

EXCEL_PATH = Path(
    "/Users/pierrelaurent/Desktop/OptiPro/Clients/SAPAL/Fournisseurs"
    "/Procity/tarifprocityvialux2026-fr.v1.7-699.xlsx"
)
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://dpycswobcixsowvxnvdc.supabase.co")
SERVICE_KEY  = os.environ.get("SUPABASE_SERVICE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRweWNzd29iY2l4c293dnhudmRjIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NDY0NjM1MywiZXhwIjoyMDkwMjIyMzUzfQ.FlZw1NjSDVR5dKE4fubr9iFzYIkvv_MqYQ6xHumX27g")
STORAGE_BUCKET = "product-images"
API_HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates",
}
SHEETS = ["MOBILIER URBAIN", "AIRES DE JEUX", "ÉQUIPEMENTS SPORTIFS", "MIROIRS"]

# Colonnes Excel (0-indexees) :
COL_REF      = 0   # REFERENCE
COL_NAME     = 4   # DESIGNATION PRODUIT
COL_FULL     = 5   # DESIGNATION COMPLETE
COL_FINITION = 6   # FINITION
COL_DIMS     = 9   # DIMENSIONS
COL_POIDS    = 10  # POIDS
COL_PRIX     = 12  # PRIX NETS
COL_COLORIS  = 13  # COLORIS
COL_DELAI    = 15  # DELAIS
COL_URL      = 17  # URL SITE PROCITY

SCRAPE_DELAY = 1.0  # secondes entre chaque requete Procity


# -- Helpers -----------------------------------------------------------------

def clean(v) -> str:
    """Nettoie une valeur Excel."""
    if v is None:
        return ""
    s = str(v).strip()
    return s if s not in ("-", "–", "—") else ""


def find_product_id_in_supabase(procity_ref: str) -> str | None:
    """Cherche l'id du produit Supabase par sa reference Procity."""
    url = (
        f"{SUPABASE_URL}/rest/v1/products"
        f"?reference=eq.{procity_ref}&select=id&limit=1"
    )
    r = requests.get(url, headers=API_HEADERS, timeout=10)
    data = r.json()
    if isinstance(data, list) and data:
        return str(data[0]["id"])
    return None


def scrape_procity_page(url: str) -> dict:
    """
    Scrappe une fiche produit procity.eu.
    Retourne:
      {
        "images": ["https://...jpg", ...],
        "options": {"Coloris": [...], "Longueur": [...], ...}
      }
    """
    result = {"images": [], "options": {}}
    try:
        resp = requests.get(
            url,
            timeout=15,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36"
                )
            },
        )
        if resp.status_code != 200:
            print(f"    [!] HTTP {resp.status_code} pour {url}")
            return result

        soup = BeautifulSoup(resp.text, "lxml")

        # -- Images --
        img_candidates = []
        for img in soup.find_all("img"):
            src = img.get("src") or img.get("data-src") or ""
            if "procity" in src and any(
                ext in src.lower()
                for ext in [".jpg", ".jpeg", ".png", ".webp"]
            ):
                if not any(
                    skip in src.lower()
                    for skip in [
                        "logo", "icon", "picto", "banner", "thumb_",
                    ]
                ):
                    if src not in img_candidates:
                        img_candidates.append(src)

        result["images"] = [
            src if src.startswith("http") else f"https://procity.eu{src}"
            for src in img_candidates
        ][:10]

        # -- Options (dropdowns / swatches) --
        for sel in soup.find_all("select"):
            label_el = sel.find_previous(["label", "span", "p"])
            label = (
                label_el.get_text(strip=True)
                if label_el
                else sel.get("name", "Option")
            )
            values = [
                opt.get_text(strip=True)
                for opt in sel.find_all("option")
                if opt.get_text(strip=True) and opt.get("value")
            ]
            if values:
                result["options"][label] = values

        for btn_group in soup.find_all(
            class_=re.compile(r"swatch|option|color|variant", re.I)
        ):
            label_el = btn_group.find_previous(["label", "span", "h3", "p"])
            label = (
                label_el.get_text(strip=True) if label_el else "Option"
            )
            values = []
            for btn in btn_group.find_all(
                ["button", "a", "span"], attrs={"data-value": True}
            ):
                v = btn.get("data-value", "").strip()
                if v:
                    values.append(v)
            if values:
                result["options"].setdefault(label, values)

    except Exception as e:
        print(f"    [!] Erreur scraping {url}: {e}")

    return result


def upload_image_to_storage(
    image_url: str, product_ref: str, idx: int
) -> str | None:
    """
    Telecharge une image et l'upload dans Supabase Storage.
    Retourne l'URL publique ou None en cas d'erreur.
    """
    try:
        resp = requests.get(
            image_url,
            timeout=15,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36"
                )
            },
        )
        if resp.status_code != 200:
            return None

        content_type = (
            resp.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()
        )
        ext = ".jpg"
        if "png" in content_type:
            ext = ".png"
        elif "webp" in content_type:
            ext = ".webp"

        url_hash = hashlib.md5(image_url.encode()).hexdigest()[:8]
        filename = f"procity/{product_ref}/{idx:02d}_{url_hash}{ext}"

        upload_url = (
            f"{SUPABASE_URL}/storage/v1/object/{STORAGE_BUCKET}/{filename}"
        )
        headers = {
            "apikey": SERVICE_KEY,
            "Authorization": f"Bearer {SERVICE_KEY}",
            "Content-Type": content_type,
            "x-upsert": "true",
        }
        up = requests.post(
            upload_url, data=resp.content, headers=headers, timeout=30
        )
        if up.status_code in (200, 201):
            return (
                f"{SUPABASE_URL}/storage/v1/object/public"
                f"/{STORAGE_BUCKET}/{filename}"
            )
        else:
            print(
                f"      [!] Upload echoue {up.status_code}: "
                f"{up.text[:100]}"
            )
            return None
    except Exception as e:
        print(f"      [!] Erreur upload {image_url}: {e}")
        return None


def delete_variants_for_product(product_id: str) -> bool:
    """Supprime les variantes existantes. Retourne True si succès."""
    try:
        r = requests.delete(
            f"{SUPABASE_URL}/rest/v1/product_variants",
            params={"product_id": f"eq.{product_id}"},
            headers=API_HEADERS,
            timeout=10,
        )
        return r.status_code in (200, 204)
    except Exception as e:
        print(f"    ⚠ Delete error: {e}")
        return False


def upsert_variants(rows: list[dict]) -> bool:
    """Insere ou met a jour des lignes dans product_variants."""
    if not rows:
        return True
    url = f"{SUPABASE_URL}/rest/v1/product_variants"
    r = requests.post(url, json=rows, headers=API_HEADERS, timeout=30)
    if r.status_code in (200, 201):
        return True
    print(f"    [x] Upsert error {r.status_code}: {r.text[:200]}")
    return False


# -- Lecture Excel -----------------------------------------------------------

def read_excel() -> dict[str, list[dict]]:
    """
    Retourne: { "206200": [ {row data}, ... ], ... }
    Une cle = une reference produit Procity.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    products: dict[str, list[dict]] = {}

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if not header_found:
                if row[COL_REF] == "REFERENCE":
                    header_found = True
                continue

            ref = clean(row[COL_REF])
            if not ref or not str(ref).replace(".", "").isdigit():
                continue

            url = clean(row[COL_URL]) if len(row) > COL_URL else ""
            if not url.startswith("http"):
                continue

            prix_raw = row[COL_PRIX] if len(row) > COL_PRIX else None
            try:
                prix = float(prix_raw) if prix_raw is not None else 0.0
            except (ValueError, TypeError):
                prix = 0.0

            entry = {
                "ref":      str(ref),
                "name":     clean(row[COL_NAME]) if len(row) > COL_NAME else "",
                "full":     clean(row[COL_FULL]) if len(row) > COL_FULL else "",
                "finition": clean(row[COL_FINITION]) if len(row) > COL_FINITION else "",
                "dims":     clean(row[COL_DIMS]) if len(row) > COL_DIMS else "",
                "poids":    clean(row[COL_POIDS]) if len(row) > COL_POIDS else "",
                "prix":     prix,
                "coloris":  clean(row[COL_COLORIS]) if len(row) > COL_COLORIS else "",
                "delai":    clean(row[COL_DELAI]) if len(row) > COL_DELAI else "",
                "url":      url,
            }
            products.setdefault(str(ref), []).append(entry)

    wb.close()
    return products


# -- Main --------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Import variantes Procity dans Supabase"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Affiche sans toucher la base",
    )
    parser.add_argument(
        "--limit", type=int, default=0,
        help="Traite seulement N produits",
    )
    parser.add_argument(
        "--product-id", type=str, default="",
        help="Traite seulement cette ref Procity",
    )
    args = parser.parse_args()

    print("Lecture Excel...")
    excel_data = read_excel()
    print(f"   {len(excel_data)} references Procity trouvees dans l'Excel")

    if args.product_id:
        excel_data = {
            k: v for k, v in excel_data.items() if k == args.product_id
        }
        if not excel_data:
            print(f"   [!] Reference {args.product_id} non trouvee dans l'Excel")
            return

    processed = 0
    skipped = 0
    total_variants = 0

    for procity_ref, rows in excel_data.items():
        if args.limit and processed >= args.limit:
            break

        url = rows[0]["url"]
        product_name = rows[0]["name"]
        print(f"\n--- [{procity_ref}] {product_name}")
        print(f"   {len(rows)} ligne(s) Excel | URL: {url}")

        # 1) Trouver l'id Supabase du produit
        supabase_product_id = find_product_id_in_supabase(procity_ref)
        if not supabase_product_id:
            print("   [!] Produit non trouve en base Supabase -- skipping")
            skipped += 1
            continue

        print(f"   OK product_id Supabase: {supabase_product_id}")

        # 2) Scraper la page Procity
        time.sleep(SCRAPE_DELAY)
        scraped = scrape_procity_page(url)
        print(
            f"   {len(scraped['images'])} image(s) trouvee(s) sur procity.eu"
        )
        if scraped["options"]:
            for k, v in scraped["options"].items():
                suffix = "..." if len(v) > 5 else ""
                print(f"      Options [{k}]: {v[:5]}{suffix}")

        # 3) Upload des images dans Supabase Storage
        stored_image_urls: list[str] = []
        if not args.dry_run:
            for i, img_url in enumerate(scraped["images"]):
                print(
                    f"   Upload image {i + 1}/{len(scraped['images'])}...",
                    end=" ",
                    flush=True,
                )
                stored = upload_image_to_storage(img_url, procity_ref, i)
                if stored:
                    stored_image_urls.append(stored)
                    print("OK")
                else:
                    print("FAIL")
        else:
            stored_image_urls = scraped["images"]

        # 4) Construire les variantes depuis l'Excel
        # Note : toutes les variantes d'un même produit partagent les mêmes images pour l'instant.
        # Le site Procity ne permet pas de corréler automatiquement les images à une combinaison
        # coloris/dimension spécifique. Amélioration future possible si Procity expose cette info.
        variant_rows = []
        for excel_row in rows:
            parts = [
                p
                for p in [
                    excel_row["dims"],
                    (
                        excel_row["coloris"]
                        if excel_row["coloris"] not in ("Standard", "")
                        else ""
                    ),
                    excel_row["finition"],
                ]
                if p
            ]
            label = (
                " — ".join(parts)
                if parts
                else (excel_row["full"] or excel_row["name"])
            )

            variant_rows.append({
                "product_id":     supabase_product_id,
                "reference":      excel_row["ref"],
                "label":          label,
                "dimensions":     excel_row["dims"],
                "finition":       excel_row["finition"],
                "coloris":        excel_row["coloris"],
                "poids":          excel_row["poids"],
                "price":          excel_row["prix"],
                "delai":          str(excel_row["delai"]),
                "specifications": {},
                "images":         stored_image_urls,
            })

        total_variants += len(variant_rows)
        print(f"   -> {len(variant_rows)} variante(s) a inserer")

        if args.dry_run:
            for v in variant_rows[:3]:
                print(
                    f"      DRY: ref={v['reference']} "
                    f"label={v['label'][:60]} "
                    f"prix={v['price']} imgs={len(v['images'])}"
                )
            if len(variant_rows) > 3:
                print(f"      ... et {len(variant_rows) - 3} autres")
        else:
            if not delete_variants_for_product(supabase_product_id):
                print("   [!] Delete échoué — skipping pour éviter les doublons")
                skipped += 1
                continue
            for i in range(0, len(variant_rows), 25):
                batch = variant_rows[i : i + 25]
                ok = upsert_variants(batch)
                status = "OK" if ok else "FAIL"
                print(
                    f"   {status} Lot {i // 25 + 1}: {len(batch)} variantes"
                )

        processed += 1

    print(f"\n{'=' * 60}")
    print(
        f"Traites : {processed} produits | "
        f"Ignores (non trouves en base) : {skipped}"
    )
    print(f"Total variantes : {total_variants}")
    if args.dry_run:
        print("(DRY RUN -- rien n'a ete modifie en base)")


if __name__ == "__main__":
    main()
