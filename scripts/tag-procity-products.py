#!/usr/bin/env python3
"""
Script one-shot : lit le tarif ProCity 2026 et génère le SQL
pour taguer les produits correspondants en base.

Usage :
  /Library/Developer/CommandLineTools/usr/bin/python3 scripts/tag-procity-products.py \
    > /tmp/tag-procity.sql

Puis exécuter le fichier SQL généré dans Supabase SQL Editor.
"""
import sys
import openpyxl

EXCEL_PATH = "Procity/tarifprocityvialux2026-fr.v1.7-699.xlsx"  # relatif à la racine du projet SAPAL

# Configuration de chaque onglet :
# sheet_name → (col_index_famille, col_index_type, col_index_reference)
# Les index sont 0-based
SHEETS = {
    "MOBILIER URBAIN":       (2, 3, 0),
    "AIRES DE JEUX":         (2, 3, 0),
    "ÉQUIPEMENTS SPORTIFS":  (2, 3, 0),
    "MIROIRS":               (3, 4, 0),  # col A = ref SPL, col D = famille, col E = type
}

HEADER_ROWS = 5  # Les 5 premières lignes sont des headers dans chaque onglet

def escape_sql(s: str) -> str:
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"

def main():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"-- ERREUR : fichier introuvable : {EXCEL_PATH}", file=sys.stderr)
        print(f"-- Lancer ce script depuis la racine du dossier client SAPAL", file=sys.stderr)
        sys.exit(1)

    # mapping : reference -> (sheet, family, type)
    mapping: dict[str, tuple[str, str, str]] = {}

    for sheet_name, (fam_idx, type_idx, ref_idx) in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"-- ATTENTION : onglet '{sheet_name}' introuvable dans le fichier", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < HEADER_ROWS:
                continue
            ref = row[ref_idx] if len(row) > ref_idx else None
            fam = row[fam_idx] if len(row) > fam_idx else None
            typ = row[type_idx] if len(row) > type_idx else None
            if ref and fam:
                ref_str = str(int(ref)) if isinstance(ref, float) else str(ref).strip()
                fam_str = str(fam).strip()
                typ_str = str(typ).strip() if typ else None
                if ref_str not in mapping:
                    mapping[ref_str] = (sheet_name, fam_str, typ_str)

    print("-- SQL généré automatiquement par tag-procity-products.py")
    print(f"-- {len(mapping)} références ProCity trouvées dans l'Excel")
    print()

    count = 0
    for ref, (sheet, family, typ) in sorted(mapping.items()):
        print(
            f"UPDATE products SET "
            f"supplier = 'procity', "
            f"procity_sheet = {escape_sql(sheet)}, "
            f"procity_family = {escape_sql(family)}, "
            f"procity_type = {escape_sql(typ)} "
            f"WHERE reference = {escape_sql(ref)};"
        )
        count += 1

    print()
    print(f"-- Total : {count} UPDATE générés")
    print()
    print("-- Vérification après UPDATE :")
    print("-- SELECT COUNT(*) FROM products WHERE supplier = 'procity';")

if __name__ == "__main__":
    main()
